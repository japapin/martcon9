import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.graph_objects as go

# Configuração da página
st.set_page_config(page_title="Análise de Estoque", layout="wide")
st.title("📈 Análise de Cobertura de Estoque")

uploaded_file = st.file_uploader("Carregue seu arquivo Excel (análise.xlsx)", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    required_cols = ["Filial", "Cobertura Atual", "Vlr Estoque Tmk", "Mercadoria", "Saldo Pedido"]
    if not all(col in df.columns for col in required_cols):
        missing_cols = [col for col in required_cols if col not in df.columns]
        st.error(f"⚠ Arquivo inválido! Faltam as colunas: {', '.join(missing_cols)}")
        st.stop()

    df = df.rename(columns={
        "Filial": "filial",
        "Cobertura Atual": "cobertura_dias",
        "Vlr Estoque Tmk": "valor_estoque",
        "Saldo Pedido": "saldo_pedido"
    })

    saldo_totais = df.groupby("filial")["saldo_pedido"].sum()

    def calcular_media_ponderada(grupo):
        grupo = grupo[grupo["valor_estoque"] > 0]
        if grupo.empty:
            return 0
        return np.average(grupo["cobertura_dias"], weights=grupo["valor_estoque"])

    def calcular_media_simples(grupo):
        return grupo["cobertura_dias"].mean()

    cobertura = (
        df.groupby("filial")
        .apply(lambda grupo: pd.Series({
            "Cobertura Média Ponderada (dias)": calcular_media_ponderada(grupo),
            "Cobertura Média Simples (dias)": calcular_media_simples(grupo)
        }))
        .round(2)
        .reset_index()
        .rename(columns={"filial": "Filial"})
    )

    cobertura["Saldo Pedido Total"] = cobertura["Filial"].map(saldo_totais)

    df['faixa'] = pd.cut(
        df['cobertura_dias'],
        bins=[-np.inf, 0, 15, 30, 45, 60, np.inf],
        labels=["<=0 dias", "1-15 dias", "16-30 dias", "31-45 dias", "46-60 dias", "Mais de 60 dias"],
        right=False
    )

    resumo_valores = df.groupby(['filial', 'faixa'])['saldo_pedido'].sum().unstack().fillna(0)
    resumo_valores['TOTAL'] = resumo_valores.sum(axis=1)
    resumo_valores = resumo_valores.reset_index()

    resumo_percentuais = resumo_valores.copy()
    for col in resumo_percentuais.columns[1:-1]:
        resumo_percentuais[col] = (resumo_percentuais[col] / resumo_percentuais['TOTAL'] * 100).round(2)
    resumo_percentuais = resumo_percentuais.drop(columns=['TOTAL'])

    # Exibição no Streamlit
    st.subheader("📌 Cobertura Média por Filial")
    st.dataframe(cobertura, use_container_width=True)

    st.subheader("📊 Distribuição por Faixa de Cobertura (Saldo de Pedido)")
    st.markdown("Valores Absolutos (R$)")
    st.dataframe(resumo_valores, use_container_width=True)
    st.markdown("Percentuais por Faixa (%)")
    st.dataframe(resumo_percentuais, use_container_width=True)

    # 🔷 Gráfico de Pareto por filial
    st.subheader("📈 Gráfico de Pareto por Filial")
    filiais = resumo_percentuais['filial'].unique()
    filial_selecionada = st.selectbox("Selecione a Filial para o Gráfico de Pareto", filiais)

    df_filial = resumo_percentuais[resumo_percentuais['filial'] == filial_selecionada].drop(columns='filial').T
    df_filial.columns = ['percentual']
    df_filial = df_filial.sort_values(by='percentual', ascending=False)
    df_filial['acumulado'] = df_filial['percentual'].cumsum()
    df_filial['acumulado_perc'] = (df_filial['acumulado'] / df_filial['percentual'].sum()) * 100

    fig = go.Figure()

    fig.add_trace(go.Bar(
        x=df_filial.index,
        y=df_filial['percentual'],
        name='Percentual por Faixa',
        marker_color='royalblue'
    ))

    fig.add_trace(go.Scatter(
        x=df_filial.index,
        y=df_filial['acumulado_perc'],
        name='Acumulado (%)',
        yaxis='y2',
        mode='lines+markers',
        marker=dict(color='crimson')
    ))

    fig.update_layout(
        title=f'📈 Gráfico de Pareto - {filial_selecionada}',
        yaxis=dict(title='Percentual Individual (%)'),
        yaxis2=dict(title='Percentual Acumulado (%)', overlaying='y', side='right'),
        xaxis=dict(title='Faixa de Cobertura'),
        legend=dict(x=0.01, y=0.99),
        bargap=0.3
    )

    st.plotly_chart(fig, use_container_width=True)

    # Geração do Excel
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Consolidado"

    def escrever_tabela_com_estilo(ws, df, titulo, linha_inicial):
        azul = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        fonte_branca = Font(color="FFFFFF", bold=True)
        fonte_negrito = Font(bold=True)
        borda = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        alinhamento = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=linha_inicial, start_column=1, end_row=linha_inicial, end_column=df.shape[1])
        cell_titulo = ws.cell(row=linha_inicial, column=1, value=titulo)
        cell_titulo.font = fonte_negrito
        linha = linha_inicial + 1

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=linha):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = borda
                cell.alignment = alinhamento
                if r_idx == linha:
                    cell.fill = azul
                    cell.font = fonte_branca

        return linha + len(df) + 1

    linha_atual = 1
    linha_atual = escrever_tabela_com_estilo(ws, cobertura, "Cobertura Média por Filial", linha_atual)
    linha_atual = escrever_tabela_com_estilo(ws, resumo_valores, "Distribuição por Faixa (Valores Absolutos)", linha_atual)
    linha_atual = escrever_tabela_com_estilo(ws, resumo_percentuais, "Distribuição por Faixa (Percentuais)", linha_atual)

    wb.save(output)
    st.download_button(
        label="📥 Baixar Relatório Excel",
        data=output.getvalue(),
        file_name="relatorio_estoque_formatado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )